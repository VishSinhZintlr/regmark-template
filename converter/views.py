from django.views import View
from django.http import HttpResponse
from .forms import DataUploadForm

from .models import DataCell
from django.shortcuts import render, redirect

from datetime import time, datetime, date as datetime_date


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from urllib.parse import urlparse

class UploadDataView(View):
    form_class = DataUploadForm
    template_name = 'upload_template.html'  # Replace with your template

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            # Process the form data
            product_id = form.cleaned_data['product_id']
            data_sheet_id = form.cleaned_data['data_sheet_id']
            table_id = form.cleaned_data['table_id']
            
            print('product_id: ', product_id)
            print('data_sheet_id: ', data_sheet_id)
            print('table_id: ', table_id)
            
            print('\n====================\n')
            
            # Process the uploaded file
            file = request.FILES['file']
            self.process_xlsx(file=file, table_id=table_id)
            
            
            
            # print(sheet)

            return HttpResponse("File successfully uploaded and processed.")  # Redirect after processing
        return render(request, self.template_name, {'form': form})
    
    def is_currency_format(self, number_format):
        currency_formats = ['$', '€', '£', '¥']  #! Add required currency symbols here
        return any(symbol in number_format for symbol in currency_formats)



    def process_xlsx(self, file, table_id, product_id, data_sheet_id):
        workbook = load_workbook(filename=file, data_only=True)
        sheet = workbook.active
        
        # Get a list of merged cell ranges
        merged_cell_ranges = list(sheet.merged_cells.ranges)
        merged_cells_info = {}

        # Create a dictionary mapping each cell to its merged range's start and end
        for mcr in merged_cell_ranges:
            start_col, start_row, end_col, end_row = mcr.bounds
            start_cell_address = f"{get_column_letter(start_col)}{start_row}"
            end_cell_address = f"{get_column_letter(end_col)}{end_row}"

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell_address = f"{get_column_letter(col)}{row}"
                    merged_cells_info[cell_address] = (start_cell_address, end_cell_address)
                    
                    
                    

        # Iterate over each row and column, checking for content
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell_address = cell.coordinate
                
                
                # Determine if the cell is part of a merged range
                is_merged_with = ""
                if cell_address in merged_cells_info:
                    start_cell, end_cell = merged_cells_info[cell_address]
                    # If the current cell is the start of a merged range, get the cell it's merged with
                    if cell_address == start_cell:
                        is_merged_with = end_cell if start_cell != end_cell else ""
                    # If the current cell is the end of a merged range, get the start cell of the merge
                    elif cell_address == end_cell:
                        is_merged_with = start_cell
                    # Otherwise, it's a middle cell in a merged range, get the end cell of the merge
                    else:
                        is_merged_with = end_cell

                
                # Determine the data type of the cell
                data_type = None
                if cell.data_type == 'n':
                    data_type = DataCell.DATA_TYPE.NUMBER
                elif cell.data_type == 's':
                    # Check if the string is a URL
                    parsed_url = urlparse(cell.value)
                    if parsed_url.scheme in ['http', 'https']:
                        data_type = DataCell.DATA_TYPE.DOCUMENT
                    else:
                        data_type = DataCell.DATA_TYPE.STRING
                elif cell.is_date:
                    # Check if the value is a date or a datetime
                    if isinstance(cell.value, datetime):
                        data_type = DataCell.DATA_TYPE.DATE_TIME
                    elif isinstance(cell.value, datetime_date):
                        data_type = DataCell.DATA_TYPE.DATE
                elif cell.data_type == 'd':
                    data_type = DataCell.DATA_TYPE.DATE
                elif cell.data_type == 't':
                    data_type = DataCell.DATA_TYPE.TIME
                else:
                    data_type = DataCell.DATA_TYPE.STRING 
                    
                
                # Determine if the cell has a currency format
                is_currency = self.is_currency_format(cell.number_format)

                # !Doesn't work as required
                # cell_color = cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else 'None'sad
                cell_color = cell.fill.fgColor
                
                
                
                # Check if cell is locked
                is_locked = cell.protection.locked
                
                
                # TODO: Find Foreign Reference IDs
                
                # Print cell details
                print('row_no:', row)
                print('col_no:', col)
                print('cell.value:', cell.value)
                print('address:', f"{table_id}.{row}.{col}")
                print('is_merged:', is_merged_with)
                print('cell_color:', cell_color)
                print('is_locked:', is_locked)
                print('data_type:', data_type)
                print('is_currency:', is_currency)
                print('\n====================\n')
                
                
                #TODO: Create a DataCell object
                
                DataCell.objects.create(
                    product_id=product_id,
                    data_sheet_id=data_sheet_id,
                    table_id=table_id,
                    cell_title=cell.value,
                    row_no=row,
                    col_no=col,
                    address=f"{table_id}.{row}.{col}",
                    is_merged=is_merged_with,
                    is_locked=is_locked,
                    data_type=data_type,
                    is_currency=is_currency,
                ) #! Add other fields as required


                


